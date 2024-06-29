from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views import View
from .models import Product, Cliente, Carrito, OrderPlaced, Payment, TCredito
from .forms import CustomerRegistrationForm, CustomerProfileForm, ProducAddForm, TCreditoForm
from django.contrib import messages
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
import openpyxl, locale, datetime
from django.contrib.auth.models import User
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.db import transaction
from openpyxl import Workbook

def home(request):
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    return render(request, "app/home.html", locals())

def about(request):
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    return render(request, "app/about.html", locals())

def contact(request):
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    return render(request, "app/contact.html", locals())

class CategoryView(View):
    def get(self, request, val):
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        product = Product.objects.filter(categoria=val)
        title = Product.objects.filter(categoria=val).values('nombre')
        return render(request, "app/category.html", locals())

class CategoryTitle(View):
    def get(self, request, val):
        product = Product.objects.filter(nombre=val)
        nombre = Product.objects.filter(categoria=product[0].categoria).values('nombre')
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, "app/category.html", locals())
    
class ProductDetail(View):
    def get(self, request, pk):
        product = Product.objects.get(pk=pk)
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, "app/productdetail.html", locals())
    
class CustomerRegistrationView(View):
    def get(self, request):
        form = CustomerRegistrationForm()
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, 'app/customerregistration.html', locals())
    def post(self, request):
        form = CustomerRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "!!Felicidades!!! Usuario registrado")
        else:
            messages.warning(request, "Error en la información")
        return render(request, 'app/customerregistration.html', locals())
    
class ProfileView(View):
    def get(self, request):
        form = CustomerProfileForm()
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, 'app/profile.html', locals())
    def post(self, request):
        form = CustomerProfileForm(request.POST)
        if form.is_valid():
            user = request.user
            nombre = form.cleaned_data['nombre']
            apellido = form.cleaned_data['apellido']
            email = form.cleaned_data['email']
            telefono = form.cleaned_data['telefono']
            direccion = form.cleaned_data['direccion']
            ciudad = form.cleaned_data['ciudad']
            region = form.cleaned_data['region']

            reg = Cliente(user=user, nombre=nombre, apellido=apellido, email=email, telefono=telefono, direccion=direccion, ciudad=ciudad, region=region)
            reg.save()
            messages.success(request, "!!Felicidades!! Perfil guardado exitosamente")
        else:
            messages.warning(request, "Información errónea")
        return render(request, 'app/profile.html', locals())

def address(request):
    add = Cliente.objects.filter(user=request.user)
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    return render(request, 'app/address.html', locals())

class updateAddress(View):
    def get(self, request, pk):
        add = Cliente.objects.get(pk=pk)
        form = CustomerProfileForm(instance=add)
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, 'app/updateAddress.html', locals())
    def post(self, request, pk):
        form = CustomerProfileForm(request.POST)
        if form.is_valid():
            add = Cliente.objects.get(pk=pk)
            add.nombre = form.cleaned_data['nombre']
            add.apellido = form.cleaned_data['apellido']
            add.telefono = form.cleaned_data['telefono']
            add.direccion = form.cleaned_data['direccion']
            add.ciudad = form.cleaned_data['ciudad']
            add.region = form.cleaned_data['region']
            add.save()
            messages.success(request, "Felicidades!!! perfil modificado exitosamente")
        else:
            messages.warning(request, "Error al modificar")
        return redirect("address")

def add_to_cart(request):
    user = request.user
    product_id = request.GET.get('prod_id')
    producto = Product.objects.get(id=product_id)
    Carrito(user=user, producto=producto).save()
    return redirect('/cart')

def show_cart(request):
    user = request.user
    cart = Carrito.objects.filter(user=user)
    amount = 0
    for p in cart:
        value = p.cantidad * p.producto.precio
        amount = amount + value
    totalamount = amount + 5000
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    return render(request, 'app/addtocart.html', locals())

class CheckoutView(View):
    def get(self, request):
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        user = request.user
        add = Cliente.objects.filter(user=user)
        cart_items = Carrito.objects.filter(user=user)
        famount = 0
        for p in cart_items:
            value = p.cantidad * p.producto.precio
            famount += value
        totalamount = famount + 5000
        request.session['totalamount'] = totalamount
        return render(request, 'app/checkout.html', locals())
    def post(self, request):
        return redirect('payment_form')
    
class PaymentFormView(View):
    def get(self, request):
        return render(request, 'app/payment_form.html')
    def post(self, request):
        full_name = request.POST.get('full_name')
        card_number = request.POST.get('card_number')
        expiry_date = request.POST.get('expiry_date')
        return redirect('paymentdone')
   
def payment_done(request):
    user = request.user
    customer = Cliente.objects.get(user=user)
    totalamount = request.session.get('totalamount', 0)
    cart = Carrito.objects.filter(user=user)

    with transaction.atomic():
        payment = Payment.objects.create(
            user=user,
            amount=totalamount,
            order_id="",
            payment_status="Pendiente",
            payment_id="",
            tipopago = "Pago con Tarjeta",
            paid=True,
        )

        order_items = []
        total_cost = 0

        for c in cart:
            if c.producto.cantidad < c.cantidad:
                transaction.set_rollback(True)
                messages.error(request, f'No hay suficiente stock para el producto {c.producto.nombre}')
                return redirect('carrito')

            item_cost = c.cantidad * c.producto.precio
            total_cost += item_cost

            order_items.append(OrderPlaced(
                user=user,
                customer=customer,
                product=c.producto,
                quantity=c.cantidad,
                payment=payment, 
                total_cost=item_cost
            ))

            producto = c.producto
            producto.cantidad -= c.cantidad
            producto.save()

        OrderPlaced.objects.bulk_create(order_items)

        cart.delete()

    return redirect('generate_order_pdf', payment_id=payment.id)

def orders(request):
    totalitem = 0
    if request.user.is_authenticated:
        totalitem = len(Carrito.objects.filter(user=request.user))
    order_placed = OrderPlaced.objects.filter(user=request.user)
    return render(request, 'app/orders.html', locals())

def plus_cart(request):
    if request.method == 'GET':
        prod_id=request.GET['prod_id']
        c = Carrito.objects.get(Q(producto=prod_id) & Q(user=request.user))
        c.cantidad+=1
        c.save()
        user = request.user
        cart = Carrito.objects.filter(user=user)
        amount = 0
        for p in cart:
            value = p.cantidad * p.producto.precio
            amount = amount + value
        totalamount = amount + 5000
        data={
            'quantity':c.cantidad,
            'amount':amount,
            'totalamount':totalamount
        }
        return JsonResponse(data)

def minus_cart(request):
    if request.method == 'GET':
        prod_id=request.GET['prod_id']
        c = Carrito.objects.get(Q(producto=prod_id) & Q(user=request.user))
        c.cantidad-=1
        c.save()
        user = request.user
        cart = Carrito.objects.filter(user=user)
        amount = 0
        for p in cart:
            value = p.cantidad * p.producto.precio
            amount = amount + value
        totalamount = amount + 5000
        data={
            'quantity':c.cantidad,
            'amount':amount,
            'totalamount':totalamount
        }
        return JsonResponse(data)

def remove_cart(request):
    if request.method == 'GET':
        prod_id=request.GET['prod_id']
        c = Carrito.objects.get(Q(producto=prod_id) & Q(user=request.user))
        c.delete()
        user = request.user
        cart = Carrito.objects.filter(user=user)
        amount = 0
        for p in cart:
            value = p.cantidad * p.producto.precio
            amount = amount + value
        totalamount = amount + 5000
        data={
            'amount':amount,
            'totalamount':totalamount
        }
        return JsonResponse(data)
    
class ProductAddView(View):
    def get(self, request):
        form = ProducAddForm()
        totalitem = 0
        if request.user.is_authenticated:
            totalitem = len(Carrito.objects.filter(user=request.user))
        return render(request, 'app/productadd.html', locals())
    def post(self, request):
        form = ProducAddForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Felicidades!! Producto guardado exitosamente")
        else:
            messages.warning(request, "Información errónea")
        return render(request, 'app/productadd.html', locals())
    
def export_users_to_excel(request):
    if not request.user.is_superuser:
        return HttpResponse("No tienes permiso para acceder a esta página.", status=403)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Usuarios"

    headers = ["Nombre", "Apellido", "RUT", "Teléfono", "Correo", "Dirección", "Ciudad", "Región", "Grupo"]
    sheet.append(headers)

    for user in User.objects.all():
        cliente = Cliente.objects.filter(user=user).first()
        if cliente:
            row = [
                cliente.nombre,
                cliente.apellido,
                user.username,
                cliente.telefono,
                cliente.email,
                cliente.direccion,
                cliente.ciudad,
                cliente.region,
                ', '.join([group.name for group in user.groups.all()])
            ]
            sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    workbook.save(response)
    return response

CATEGORY_CHOICES=(
    ('JO','Joyería de autor'),
    ('CA','Carteras'),
    ('VE','Vestimenta'),
    ('CZ','Calzado')
)

CATEGORY_DICT = dict(CATEGORY_CHOICES)

def export_products_to_excel(request):
    if not request.user.is_superuser:
        return HttpResponse("No tienes permiso para acceder a esta página.", status=403)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Productos"

    headers = ["Nombre", "Precio", "Cantidad", "SKU", "Categoría"]
    sheet.append(headers)

    for prod in Product.objects.all():
        if prod:
            row = [
                prod.nombre,
                prod.precio,
                prod.cantidad,
                prod.sku,
                CATEGORY_DICT.get(prod.categoria, "Desconocido"),
            ]
            sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    workbook.save(response)
    return response

def generate_order_pdf(request, payment_id):
    user = request.user
    cliente = Cliente.objects.get(user=user)
    orders = OrderPlaced.objects.filter(user=user, payment_id=payment_id)
    locale.setlocale(locale.LC_ALL, 'es_CL.UTF-8')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="boleta_{payment_id}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 50, "Boleta")

    p.setFont("Helvetica", 12)
    p.drawString(100, height - 80, f"Nombre: {cliente.nombre} {cliente.apellido}")
    p.drawString(100, height - 100, f"RUT: {user.username}")
    p.drawString(100, height - 120, f"Email: {cliente.email}")
    p.drawString(100, height - 140, f"Método de Pago: Pago con Tarjeta")
    
    y = height - 210
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, y, "Producto")
    p.drawString(300, y, "Cantidad")
    p.drawString(400, y, "Costo de Envío")
    p.drawString(500, y, "Costo")
    p.drawString(400, height - 600, "Total")

    y -= 30
    p.setFont("Helvetica", 12)
    total_general = 0
    for order in orders:
        p.drawString(100, y, order.product.nombre)
        p.drawString(300, y, str(order.quantity))

        formatted_envio = locale.format_string("%d", order.envio, grouping=True)
        p.drawString(400, y, f"${formatted_envio}")

        formatted_cost = locale.format_string("%d", order.total_cost, grouping=True)
        p.drawString(500, y, f"${formatted_cost}")

        total = order.envio + order.total_cost
        formatted_total = locale.format_string("%d", total, grouping=True)
        p.drawString(450, height - 600, f"${formatted_total}")
        total_general += total
        y -= 20
        if y < 50:
            p.showPage()
            y = height - 50

    p.setFont("Helvetica-Bold", 12)
    formatted_total_general = locale.format_string("%d", total_general, grouping=True)
    p.drawString(400, height - 630, f"Total General: ${formatted_total_general}")

    p.showPage()
    p.save()

    return response

def export_sales_to_excel(request):
    today = datetime.date.today()

    orders = OrderPlaced.objects.filter(ordered_date__date=today)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas del Día"

    headers = ['Fecha', 'Nombre del Producto', 'Precio', 'Cantidad', 'Costo Total']
    ws.append(headers)

    for order in orders:
        row = [
            today,
            order.product.nombre,
            order.product.precio,
            order.quantity,
            order.total_cost,
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Ventas_Dia.xlsx'
    wb.save(response)

    return response